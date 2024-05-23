import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import hashlib
import json
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import requests
import time
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
import urllib3
from datetime import datetime, timedelta
from tkinter import messagebox

# 加载用户信息和充值码
def load_data():
    try:
        with open('users.json', 'r') as f:
            users = json.load(f)
    except FileNotFoundError:
        users = {}
    try:
        with open('recharge_codes.json', 'r') as f:
            recharge_codes = json.load(f)
    except FileNotFoundError:
        recharge_codes = {}
    return users, recharge_codes

def authenticate(username, password):
    for user_id, user_info in users.items():
        if user_info['username'] == username and user_info['password'] == password:
            return user_id
    return None

def recharge_points(user_id, recharge_code):
    if recharge_code in recharge_codes and not recharge_codes[recharge_code]['used']:
        if recharge_codes[recharge_code]['user_id'] == user_id:
            users[user_id]['points'] += recharge_codes[recharge_code]['points']
            recharge_codes[recharge_code]['used'] = True
            save_data()
            return True
    return False

def check_points(user_id):
    return users[user_id]['points']

def save_data():
    with open('users.json', 'w') as f:
        json.dump(users, f)
    with open('recharge_codes.json', 'w') as f:
        json.dump(recharge_codes, f)

def load_login_info():
    try:
        with open('login_info.json', 'r') as f:
            login_info = json.load(f)
    except FileNotFoundError:
        login_info = {}
    return login_info

def save_login_info(login_info):
    with open('login_info.json', 'w') as f:
        json.dump(login_info, f)

def login():
    username = username_entry.get()
    password = password_entry.get()
    user_id = authenticate(username, password)
    if user_id:
        if remember_var.get():
            login_info['username'] = username
            login_info['password'] = password
            save_login_info(login_info)
        else:
            login_info.clear()
            save_login_info(login_info)
        open_main_window(user_id)
    else:
        messagebox.showerror("错误", "无效的用户名或密码")

def open_main_window(user_id):
    login_window.destroy()

    def recharge():
        recharge_code = recharge_entry.get()
        if recharge_points(user_id, recharge_code):
            messagebox.showinfo("成功", "充值成功!")
        else:
            messagebox.showerror("错误", "无效或已使用的充值密文。")

    def check():
        points = check_points(user_id)
        messagebox.showinfo("当前积分", f"你当前的积分: {points}")

    def run_main_program():
        main_window.destroy()

        def get_time():
            return str(int(time.time() * 1000))

        root = Tk()
        root.withdraw()

        # 弹窗供用户选择Excel文件路径
        excel_file = askopenfilename(title="选择Excel文件", filetypes=[("Excel文件", "*.xlsx")])
        print("=========================欢迎使用自动参数查询系统，正在执行以下程序==========================")
        print("=========================欢迎使用自动参数查询系统，正在执行以下程序==========================")
        print("=========================欢迎使用自动参数查询系统，正在执行以下程序==========================")

        def check_cookie_file():
            if os.path.exists('手机号cookie.txt'):
                # 获取文件的修改时间
                modified_time = os.path.getmtime('手机号cookie.txt')
                # 当前时间
                current_time = time.time()
                # 检查文件是否在一小时内修改过
                if current_time - modified_time < 3600:
                    return True
            return False

        # 弹窗输入 Cookie
        def input_cookie():
            cookie = input("请输入 互动营销的Cookie：")
            # 将 Cookie 写入文件
            with open('手机号cookie.txt', 'w') as file:
                file.write(cookie)

        # 获取保存的 Cookie
        def get_cookie():
            with open('手机号cookie.txt', 'r') as file:
                cookie = file.read()
            return cookie

        # 检查是否需要重新输入 Cookie
        if check_cookie_file():
            cookie = get_cookie()
        else:
            input_cookie()
            cookie = get_cookie()

        # 使用openpyxl库加载Excel文件
        workbook = openpyxl.load_workbook(excel_file)

        # 选择默认的工作表
        sheet = workbook.active

        # 创建一个空列表来存储"A"列中的值
        n = []

        # 循环读取"A"列中的值
        for cell in sheet['A'][1:]:
            value = cell.value
            if value is None:
                break
                # 这里处理每个单元格的值，例如，打印出来
                print(value)
                # 每次循环减少 1 个积分
                if users[user_id]['points'] > 0:
                    users[user_id]['points'] -= 1
                    save_data()
                    time.sleep(1)  # 模拟每次循环的延迟
                    if users[user_id]['points'] == 0:
                        messagebox.showinfo("积分不足", "你的积分已用完，主程序停止运行。")
                        break
                else:
                    messagebox.showinfo("积分不足", "你的积分已用完，主程序停止运行。")
                    break
        # 构建查询参数和请求头

        headers = {
            "Accept": 'application/json, text/javascript, */*; q=0.01',
            "Accept-Encoding": 'gzip, deflate',
            "Accept-Language": 'zh-CN,zh;q=0.9',
            "Connection": 'keep-alive',
            "Content-Length": '177',
            "Content-Type": 'application/x-www-form-urlencoded; charset=UTF-8',
            "Cookie": cookie,
            "DNT": '1',
            "Host": '10.248.250.254:18801',
            "Origin": 'http://10.248.250.254:18801',
            "Referer": 'http://10.248.250.254:18801/system/cou_query_analysis/user_coupon;jsessionid=88BCA04998DF5C98841ED6FBC496E6C7',
            "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36 QIHU 360ENT',
            "X-Requested-With": 'XMLHttpRequest'

        }

        url = 'http://10.248.250.254:18801/system/jqGrid/cou_query_analysis/getUserCouponList.action?randomTimeTemp=' + get_time()

        # 创建一个空DataFrame来存储结果
        result_df = pd.DataFrame(columns=['userId', 'phone'])

        # 循环查询并处理每个值
        for value in n:
            print('正在查询：===========================', value, '的手机号码，请稍等！')

            # 构建查询参数
            data = {
                "_search": 'true',
                "nd": get_time(),
                "rows": '10',
                "page": '1',
                "sidx": 'id',
                "sord": 'desc',
                "phone": '',
                "couTypeCode": '',
                "userId": value,
                "couDistUsedUser.mktActivityId": '',
                "state": '',
                "couNo": '',
                "distOrderNo": '',
                "packageNo": '',
                "distTime": ''
            }
            query_string = '&'.join([f"{key}={value}" for key, value in data.items()])

            # 发送查询请求
            res = requests.post(url=url, data=query_string, headers=headers)

            # 解析查询结果
            soup = BeautifulSoup(res.content, 'html.parser')
            result = json.loads(str(soup))

            # 假设data是包含字典的列表
            data = result['data']
            user_ids = [item['userId'] for item in data]
            phones = [item['phone'] for item in data]

            # 将查询结果添加到DataFrame，并进行去重
            temp_df = pd.DataFrame({'userId': user_ids, 'phone': phones})
            temp_df = temp_df.drop_duplicates()
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

        # 写入数据到Excel文件
        sheet.cell(row=1, column=2, value='userId')
        sheet.cell(row=1, column=3, value='phone')
        for index, row in result_df.iterrows():
            sheet.cell(row=index + 2, column=2, value=row['userId'])  # 将userId写入B列
            sheet.cell(row=index + 2, column=3, value=row['phone'])  # 将phone写入C列

        # 保存Excel文件
        workbook.save(excel_file)
        workbook.close()
        # print(result_df)
        print("=========================手机号查询完毕，进入参数查询阶段==========================")
        print("=========================手机号查询完毕，进入参数查询阶段==========================")
        print("=========================手机号查询完毕，进入参数查询阶段==========================")

        # 以下为参数查询
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        current_time = datetime.now()
        root.withdraw()
        new_time = current_time - timedelta(days=4)
        formatted_time = new_time.strftime('%Y-%m-%d %H:%M:%S')
        print(formatted_time)
        formatted_time1 = current_time.strftime('%Y-%m-%d %H:%M:%S')
        # 将指定时间改为当前时间
        date_string1 = formatted_time1
        print(date_string1)

        def check_cookie_file():
            if os.path.exists('云钱包cookie.txt'):
                # 获取文件的修改时间
                modified_time = os.path.getmtime('云钱包cookie.txt')
                # 当前时间
                current_time = time.time()
                # 检查文件是否在一小时内修改过
                if current_time - modified_time < 3600:
                    return True
            return False

        # 弹窗输入 Cookie
        def input_cookie():
            cookie = input("请输入 云钱包的Cookie：")
            # 将 Cookie 写入文件
            with open('云钱包cookie.txt', 'w') as file:
                file.write(cookie)

        # 获取保存的 Cookie
        def get_cookie():
            with open('云钱包cookie.txt', 'r') as file:
                cookie = file.read()
            return cookie

        # 检查是否需要重新输入 Cookie
        if check_cookie_file():
            cookie = get_cookie()
        else:
            input_cookie()
            cookie = get_cookie()

        # 使用openpyxl库加载Excel文件
        workbook = openpyxl.load_workbook(excel_file)
        # 选择默认的工作表
        sheet = workbook.active
        # 创建一个空列表来存储"C"列中的值
        nb = []
        # 循环读取"A"列中的值
        for cell in sheet['C'][1:]:
            value = cell.value
            if value is None:
                break
            nb.append(value)

        # 构建查询参数1和请求头
        headers = {
            "Accept": 'application/json, text/javascript, */*; q=0.01',
            "Accept-Language": 'zh-Hans-CN, zh-Hans; q=0.8, en-US; q=0.6, en-GB; q=0.4, en; q=0.2',
            "Connection": 'Keep-Alive',
            "Content-Length": '55',
            "Content-Type": 'application/x-www-form-urlencoded; charset=UTF-8',
            "Cookie": cookie,
            "sec-fetch-dest": 'empty',
            "sec-fetch-mode": 'cors',
            "sec-fetch-site": 'same-origin',
            "Referer": 'https://ewcc.sinopec.com/acc/html/acc/accCsrInfoList.html',
            "User-Agent": 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729;  QIHU 360ENT)',
            "X-Requested-With": 'XMLHttpRequest'
        }
        # 参数2的请求
        headers1 = {
            "Accept": 'application/json, text/javascript, */*; q=0.01',
            "Accept-Language": 'zh-Hans-CN, zh-Hans; q=0.8, en-US; q=0.6, en-GB; q=0.4, en; q=0.2',
            "Connection": 'Keep-Alive',
            "Content-Length": '55',
            "Content-Type": 'application/x-www-form-urlencoded; charset=UTF-8',
            "Cookie": cookie,
            "sec-fetch-dest": 'empty',
            "sec-fetch-mode": 'cors',
            "sec-fetch-site": 'same-origin',
            "Referer": 'https://ewcc.sinopec.com/acc/html/acc/accAccountDetail.html',
            "User-Agent": 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729;  QIHU 360ENT)',
            "X-Requested-With": 'XMLHttpRequest'
        }

        # 查询1参数地址
        url = 'https://ewcc.sinopec.com/acc/acc/getAccCsrInfo'
        # 查询3参数地址
        url3 = 'https://ewcc.sinopec.com/acc/acc/getAccountDetailById'
        # 查询2参数地址
        url2 = 'https://ewcc.sinopec.com/acc/acc/getAccInfoByCsrid'
        data_storage = {}
        # 循环查询手机号的第一个参数
        names = []
        csrids = []
        accids = []

        for i, value in enumerate(nb, start=2):
            # 构建查询参数
            print('正在查询：===========================', value, '的石化钱包参数，请稍等！')
            data = {
                "rows": 50,
                "page": 1,
                "searchType": 'phone',
                "searchValue": value
            }

            # 发送查询请求
            res = requests.post(url=url, data=data, headers=headers, verify=False)

            time.sleep(0.2)
            if res.status_code == 200:
                response_text = res.text
                data1 = json.loads(response_text)
                # print(data1)

                if isinstance(data1, dict) and data1.get('data', {}).get('list'):
                    first_item = data1['data']['list'][0]
                    name = first_item['name']

                    # 检查isDelete键的值
                    if 'isDelete' in first_item and not first_item['isDelete']:
                        csrid = first_item['csrid']
                    else:
                        csrid = ''
                else:
                    name = ''
                    csrid = ''

                    print("==========以上这一行的手机号查询参数结果为空，可以删除===========")
                    # 数据清洗
                    names.append("")
                    csrids.append("")
                    accids.append("")
                    continue

                names.append(name)
                csrids.append(csrid)

                # 将 res 的值作为字符串写入到底列
                df = pd.DataFrame({'Column4': [res.text]})
                for cell, value in zip(sheet.iter_rows(min_row=i, min_col=4, max_col=4), df['Column4']):
                    cell[0].value = value

                if not all(csrid):
                    print('csrids 中存在空值，跳过发送请求')
                    continue

                data2 = {
                    "csrid": csrid
                }

                query_string = '&'.join([f"{key}={value}" for key, value in data2.items()])

                # 发送查询请求
                res2 = requests.post(url=url2, data=data2, headers=headers1, verify=False)
                if res2.status_code == 200:
                    dat = res2.json()
                    if isinstance(dat, dict):
                        accid = dat['data']['accid']
                    else:
                        accid = [item['accid'] for item in dat]
                    accids.append(accid)
                    # print(accid)
                else:
                    print("请求失败:", res.status_code)

                if len(csrids) > 0 and len(names) > 0 and len(accids) > 0:
                    if i >= 2 and i - 2 < len(csrids) and i - 2 < len(names) and i - 2 < len(accids):
                        data3 = {
                            "rows": 50,
                            "page": 1,
                            "csrid": csrids[i - 2],
                            "csrname": names[i - 2],
                            "accid": accids[i - 2],
                            "tradetype": 'pay',
                            "tranTimeBegin": formatted_time,
                            "tranTimeEnd": date_string1
                        }

                        query_string = '&'.join([f"{key}={value}" for key, value in data3.items()])

                        # 发送查询请求
                        res3 = requests.post(url=url3, data=data3, headers=headers1, verify=False)
                        df1 = pd.DataFrame({'Column5': [res3.text]})
                        for cell, value in zip(sheet.iter_rows(min_row=i, min_col=5, max_col=5), df1['Column5']):
                            cell[0].value = value
                        # print(res3)

        # 删除第一行
        # sheet.delete_rows(1)

        # 删除第一列、第二列
        sheet.delete_cols(1, 2)
        workbook.save(excel_file)
        workbook.close()
        print("===========所有参数查询已完成，谢谢使用本查询系统===========")
        print("===========所有参数查询已完成，谢谢使用本查询系统===========")
        messagebox.showinfo("查询完毕", "查询已完成，请自行打开文件查看结果。")

    main_window = tk.Tk()
    main_window.title("用户界面")

    main_frame = ttk.Frame(main_window, padding="20 20 20 20")
    main_frame.pack(fill=tk.BOTH, expand=True)

    recharge_label = ttk.Label(main_frame, text="请输入充值密文:")
    recharge_label.grid(column=0, row=0, sticky=tk.W, pady=10)

    recharge_entry = ttk.Entry(main_frame, width=30)
    recharge_entry.grid(column=1, row=0, pady=10)

    recharge_button = ttk.Button(main_frame, text="充值积分", command=recharge)
    recharge_button.grid(column=0, row=1, columnspan=2, pady=10)

    check_button = ttk.Button(main_frame, text="查询积分", command=check)
    check_button.grid(column=0, row=2, columnspan=2, pady=10)

    run_button = ttk.Button(main_frame, text="运行主程序", command=run_main_program)
    run_button.grid(column=0, row=3, columnspan=2, pady=10)

    exit_button = ttk.Button(main_frame, text="退出", command=main_window.destroy)
    exit_button.grid(column=0, row=4, columnspan=2, pady=10)

    for child in main_frame.winfo_children():
        child.grid_configure(padx=10, pady=5)

    main_window.mainloop()

users, recharge_codes = load_data()
login_info = load_login_info()

# 创建登录窗口
login_window = tk.Tk()
login_window.title("登录")

login_frame = ttk.Frame(login_window, padding="20 20 20 20")
login_frame.pack(fill=tk.BOTH, expand=True)

username_label = ttk.Label(login_frame, text="用户名:")
username_label.grid(column=0, row=0, sticky=tk.W, pady=10)

username_entry = ttk.Entry(login_frame, width=30)
username_entry.grid(column=1, row=0, pady=10)

password_label = ttk.Label(login_frame, text="密码:")
password_label.grid(column=0, row=1, sticky=tk.W, pady=10)

password_entry = ttk.Entry(login_frame, show='*', width=30)
password_entry.grid(column=1, row=1, pady=10)

if 'username' in login_info and 'password' in login_info:
    username_entry.insert(0, login_info['username'])
    password_entry.insert(0, login_info['password'])

remember_var = tk.BooleanVar()
remember_check = ttk.Checkbutton(login_frame, text="记住密码", variable=remember_var)
remember_check.grid(column=0, row=2, columnspan=2, pady=10)

login_button = ttk.Button(login_frame, text="登录", command=login)
login_button.grid(column=0, row=3, columnspan=2, pady=10)

for child in login_frame.winfo_children():
    child.grid_configure(padx=10, pady=5)

login_window.mainloop()
